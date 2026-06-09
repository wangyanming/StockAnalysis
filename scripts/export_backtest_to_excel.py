#!/usr/bin/env python3
"""导出 v5_build_candidate_pool 回测结果到 Excel"""

import sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from utils.dao import get_db
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BATCH = "v5_build_candidate_pool"
OUTPUT = f"/Users/wangyanming/workspace/StockAnalysis/回测结果_{BATCH}.xlsx"

db = get_db()

# 拉取数据
rows = db.fetchall(
    'SELECT * FROM backtest_results WHERE run_batch=%s ORDER BY trade_date DESC, group_rank',
    (BATCH,)
)
print(f"v5_build_candidate_pool 共 {len(rows)} 条记录")

wb = Workbook()

# 样式
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="4472C4")
yellow_fill = PatternFill("solid", fgColor="FFF2CC")
green_fill = PatternFill("solid", fgColor="D5F5E3")
red_font = Font(color="CC0000")
green_font = Font(color="006600")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def style_header(ws, headers):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

# ============== Sheet1: 全量明细 ==============
ws1 = wb.active
ws1.title = "全量明细"

cols_desc = db.fetchall('SHOW COLUMNS FROM backtest_results')
field_names = [c['Field'] for c in cols_desc if c['Field'] not in ('id', 'created_at', 'run_batch')]

style_header(ws1, field_names)

for row_idx, r in enumerate(rows, 2):
    for col_idx, fn in enumerate(field_names, 1):
        val = r.get(fn, "")
        if val is None:
            val = ""
        cell = ws1.cell(row=row_idx, column=col_idx, value=val)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
        # 条件染色：next_change 正负
        if fn == 'next_change':
            try:
                v = float(val)
                if v > 0:
                    cell.font = green_font
            except:
                pass

# 列宽
col_widths_map = {
    'trade_date': 12, 'code': 12, 'name': 16, 'source': 14,
    'group_rank': 10, 'total_score': 12, 'is_pick': 8,
    'next_open': 12, 'next_close': 12, 'next_change': 12, 'sh_change': 12,
}
for i, fn in enumerate(field_names, 1):
    ws1.column_dimensions[get_column_letter(i)].width = col_widths_map.get(fn, 14)

ws1.auto_filter.ref = ws1.dimensions
ws1.freeze_panes = 'A2'

# ============== Sheet2: 精选(is_pick=1) ==============
ws2 = wb.create_sheet("精选记录")
pick_rows = [r for r in rows if r.get('is_pick') == 1]
print(f"   精选记录: {len(pick_rows)} 条")

style_header(ws2, field_names)
for row_idx, r in enumerate(pick_rows, 2):
    for col_idx, fn in enumerate(field_names, 1):
        val = r.get(fn, "")
        if val is None: val = ""
        cell = ws2.cell(row=row_idx, column=col_idx, value=val)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
        if fn == 'next_change':
            try:
                v = float(val)
                if v > 0: cell.font = green_font
            except: pass

for i, fn in enumerate(field_names, 1):
    ws2.column_dimensions[get_column_letter(i)].width = col_widths_map.get(fn, 14)
ws2.auto_filter.ref = ws2.dimensions
ws2.freeze_panes = 'A2'

# ============== Sheet3: 汇总统计 ==============
ws3 = wb.create_sheet("汇总统计")

# 统计
source_groups = {}
for r in rows:
    src = r.get('source', '未知')
    if src not in source_groups:
        source_groups[src] = {'count': 0, 'up': 0, 'down': 0, 'zero': 0, 'sum_change': 0.0}
    sg = source_groups[src]
    sg['count'] += 1
    nc = r.get('next_change')
    if nc is not None:
        nc = float(nc)
        sg['sum_change'] += nc
        if nc > 0: sg['up'] += 1
        elif nc < 0: sg['down'] += 1
        else: sg['zero'] += 1

# 大盘安全垫分析
has_sh = [r for r in rows if r.get('sh_change') is not None]
sh_up_days = [r for r in has_sh if float(r['sh_change']) >= 0]

style_header(ws3, ['指标', '值'])
summary_data = [
    ("回测批次", BATCH),
    ("总记录数", len(rows)),
    ("精选记录(is_pick=1)", len(pick_rows)),
    ("数据周期", f"{rows[-1]['trade_date']} ~ {rows[0]['trade_date']}" if rows else "无"),
    ("", ""),
]
for src, sg in sorted(source_groups.items()):
    win_rate = round(sg['up'] / sg['count'] * 100, 1) if sg['count'] else 0
    avg_chg = round(sg['sum_change'] / sg['count'], 2) if sg['count'] else 0
    summary_data.extend([
        (f"来源: {src}", ""),
        (f"  · 总次数", sg['count']),
        (f"  · 上涨日", sg['up']),
        (f"  · 下跌日", sg['down']),
        (f"  · 胜率", f"{win_rate}%"),
        (f"  · 平均次日涨幅", f"{avg_chg}%"),
        ("", ""),
    ])

summary_data.extend([
    ("大盘环境", ""),
    (f"  · 大盘上涨日选股", len(sh_up_days)),
    (f"  · 大盘下跌日选股", len(has_sh) - len(sh_up_days) if has_sh else 0),
])

for row_idx, (k, v) in enumerate(summary_data, 2):
    cell_k = ws3.cell(row=row_idx, column=1, value=str(k))
    cell_v = ws3.cell(row=row_idx, column=2, value=str(v))
    cell_k.border = thin_border
    cell_v.border = thin_border
    if k and not k.startswith('  ·') and k != '':
        cell_k.font = Font(bold=True)

ws3.column_dimensions['A'].width = 30
ws3.column_dimensions['B'].width = 20

# ============== Sheet4: 每日TOP5得分 ==============
ws4 = wb.create_sheet("每日TOP5得分")
daily_headers = ['trade_date', 'code', 'name', 'source', 'group_rank', 'total_score', 'score_chip', 'score_money', 'score_sector', 'score_trend', 'score_market', 'score_pos', 'next_change', 'sh_change']
style_header(ws4, daily_headers)

# 按日期+rank排序展示
daily_rows = sorted(rows, key=lambda r: (r['trade_date'], r.get('group_rank', 99) or 99))
for row_idx, r in enumerate(daily_rows, 2):
    for col_idx, fn in enumerate(daily_headers, 1):
        val = r.get(fn, "")
        if val is None: val = ""
        cell = ws4.cell(row=row_idx, column=col_idx, value=val)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
        if fn == 'next_change':
            try:
                v = float(val)
                if v > 0: cell.font = green_font
            except: pass
        # 高亮 is_pick=1
        if r.get('is_pick') == 1 and fn == 'group_rank':
            cell.fill = yellow_fill

for i, fn in enumerate(daily_headers, 1):
    ws4.column_dimensions[get_column_letter(i)].width = col_widths_map.get(fn, 12)
ws4.auto_filter.ref = ws4.dimensions
ws4.freeze_panes = 'A2'

wb.save(OUTPUT)
print(f"\n✅ 已导出: {OUTPUT}")
print(f"   Sheet1: 全量明细 ({len(rows)}条)")
print(f"   Sheet2: 精选记录 ({len(pick_rows)}条)")
print(f"   Sheet3: 汇总统计")
print(f"   Sheet4: 每日TOP5得分")
